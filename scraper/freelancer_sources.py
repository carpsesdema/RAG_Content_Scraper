# scraper/freelancer_sources.py

import requests
from bs4 import BeautifulSoup
from typing import List, Dict
import re
import time
import json
from urllib.parse import urljoin, urlparse
from .parser import extract_code as extract_code_from_html_markdown


class FreelancerPythonSources:
    """Scrape sources particularly valuable for freelance Python work."""

    def __init__(self, user_agent: str, timeout: int = 15):
        self.user_agent = user_agent
        self.timeout = timeout
        self.session = requests.Session()
        self.session.headers.update({'User-Agent': user_agent})

        # Common freelance tech stack
        self.freelance_keywords = {
            'web_frameworks': ['fastapi', 'django', 'flask', 'starlette'],
            'databases': ['postgresql', 'mongodb', 'sqlite', 'redis'],
            'apis': ['rest', 'graphql', 'oauth', 'jwt', 'api'],
            'payments': ['stripe', 'paypal', 'payment', 'subscription'],
            'communication': ['twilio', 'sendgrid', 'slack', 'discord', 'email'],
            'cloud': ['aws', 'docker', 'kubernetes', 'heroku', 'digital ocean'],
            'data': ['pandas', 'numpy', 'csv', 'excel', 'etl', 'pipeline'],
            'automation': ['selenium', 'schedule', 'cron', 'scraping', 'bot'],
            'testing': ['pytest', 'unittest', 'integration', 'e2e', 'mock'],
            'monitoring': ['logging', 'sentry', 'prometheus', 'health check']
        }

    def fetch_python_patterns(self, query: str, logger) -> List[str]:
        """Fetch Python design patterns and best practices."""
        logger.info(f"Fetching Python patterns for: {query}")
        snippets = []

        pattern_sources = [
            {
                'url': 'https://python-patterns.guide/gang-of-four/',
                'selectors': ['pre', 'code.highlight']
            },
            {
                'url': 'https://refactoring.guru/design-patterns/python',
                'selectors': ['pre.highlight', 'code.language-python']
            }
        ]

        for source in pattern_sources:
            try:
                response = self.session.get(source['url'], timeout=self.timeout)
                response.raise_for_status()

                soup = BeautifulSoup(response.text, 'html.parser')

                for selector in source['selectors']:
                    for code_block in soup.select(selector):
                        code_text = code_block.get_text().strip()
                        if len(code_text) > 50 and ('class' in code_text or 'def' in code_text):
                            snippets.append(code_text)

                time.sleep(1)  # Be respectful

            except Exception as e:
                logger.warning(f"Error fetching patterns from {source['url']}: {e}")

        return snippets

    def fetch_fastapi_examples(self, query: str, logger) -> List[str]:
        """Fetch FastAPI examples - crucial for modern Python freelancing."""
        logger.info(f"Fetching FastAPI examples for: {query}")
        snippets = []

        fastapi_endpoints = [
            'https://fastapi.tiangolo.com/tutorial/first-steps/',
            'https://fastapi.tiangolo.com/tutorial/path-params/',
            'https://fastapi.tiangolo.com/tutorial/query-params/',
            'https://fastapi.tiangolo.com/tutorial/request-body/',
            'https://fastapi.tiangolo.com/tutorial/dependencies/',
            'https://fastapi.tiangolo.com/tutorial/security/',
            'https://fastapi.tiangolo.com/advanced/middleware/'
        ]

        for url in fastapi_endpoints:
            try:
                response = self.session.get(url, timeout=self.timeout)
                response.raise_for_status()

                soup = BeautifulSoup(response.text, 'html.parser')

                # Look for code blocks
                for code_block in soup.find_all(['pre', 'code']):
                    code_text = code_block.get_text().strip()
                    if 'fastapi' in code_text.lower() and len(code_text) > 30:
                        snippets.append(code_text)

                time.sleep(0.5)  # Be respectful

            except Exception as e:
                logger.debug(f"Error fetching FastAPI example from {url}: {e}")

        return snippets

    def fetch_automation_scripts(self, query: str, logger) -> List[str]:
        """Fetch automation and utility scripts."""
        logger.info(f"Fetching automation scripts for: {query}")
        snippets = []

        # Check if query is automation-related
        automation_terms = ['automat', 'script', 'util', 'tool', 'schedul', 'cron', 'selenium']
        if not any(term in query.lower() for term in automation_terms):
            return snippets

        # Automation script examples
        automation_examples = {
            'file_automation': '''
import os
import shutil
from pathlib import Path

def organize_downloads():
    """Organize downloads folder by file type."""
    downloads_path = Path.home() / "Downloads"

    file_types = {
        'images': ['.jpg', '.jpeg', '.png', '.gif'],
        'documents': ['.pdf', '.doc', '.docx', '.txt'],
        'videos': ['.mp4', '.avi', '.mkv'],
        'archives': ['.zip', '.rar', '.7z']
    }

    for file_type, extensions in file_types.items():
        folder_path = downloads_path / file_type
        folder_path.mkdir(exist_ok=True)

        for file_path in downloads_path.iterdir():
            if file_path.suffix.lower() in extensions:
                shutil.move(str(file_path), str(folder_path / file_path.name))
            ''',

            'email_automation': '''
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import schedule
import time

def send_automated_email(to_email, subject, body):
    """Send automated email with SMTP."""
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    sender_email = "your_email@gmail.com"
    sender_password = "your_app_password"

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = to_email
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'plain'))

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(msg)

# Schedule email
schedule.every().monday.at("09:00").do(
    send_automated_email, 
    "client@example.com", 
    "Weekly Report", 
    "Here's your weekly report..."
)
            ''',

            'web_scraping_automation': '''
import requests
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
import time

class WebScraper:
    def __init__(self, headless=True):
        options = webdriver.ChromeOptions()
        if headless:
            options.add_argument('--headless')
        self.driver = webdriver.Chrome(options=options)

    def scrape_dynamic_content(self, url, selector):
        """Scrape content from JavaScript-heavy sites."""
        self.driver.get(url)
        time.sleep(3)  # Wait for JS to load

        elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
        data = [element.text for element in elements]

        return data

    def scrape_table_data(self, url):
        """Scrape table data and convert to DataFrame."""
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')

        tables = soup.find_all('table')
        dataframes = []

        for table in tables:
            df = pd.read_html(str(table))[0]
            dataframes.append(df)

        return dataframes

    def close(self):
        self.driver.quit()
            '''
        }

        # Add relevant examples based on query
        for script_type, code in automation_examples.items():
            if any(term in query.lower() for term in script_type.split('_')):
                snippets.append(code.strip())

        return snippets

    def fetch_testing_patterns(self, query: str, logger) -> List[str]:
        """Fetch testing patterns and pytest examples."""
        logger.info(f"Fetching testing patterns for: {query}")
        snippets = []

        if 'test' not in query.lower():
            return snippets

        testing_examples = {
            'pytest_fixtures': '''
import pytest
from unittest.mock import Mock, patch

@pytest.fixture
def sample_data():
    """Provide sample data for tests."""
    return {
        'users': [
            {'id': 1, 'name': 'John Doe', 'email': 'john@example.com'},
            {'id': 2, 'name': 'Jane Smith', 'email': 'jane@example.com'}
        ]
    }

@pytest.fixture
def mock_database():
    """Mock database connection."""
    with patch('app.database.get_connection') as mock_conn:
        mock_conn.return_value = Mock()
        yield mock_conn

def test_user_creation(sample_data, mock_database):
    """Test user creation with fixtures."""
    from app.services import UserService

    service = UserService()
    user_data = sample_data['users'][0]

    result = service.create_user(user_data)

    assert result['success'] is True
    assert result['user']['name'] == user_data['name']
            ''',

            'api_testing': '''
import pytest
import requests
from fastapi.testclient import TestClient
from app.main import app

client = TestClient(app)

class TestAPI:
    def test_health_check(self):
        """Test API health check endpoint."""
        response = client.get("/health")
        assert response.status_code == 200
        assert response.json() == {"status": "healthy"}

    def test_create_user(self):
        """Test user creation endpoint."""
        user_data = {
            "name": "Test User",
            "email": "test@example.com",
            "password": "secure_password"
        }

        response = client.post("/users/", json=user_data)
        assert response.status_code == 201

        created_user = response.json()
        assert created_user["name"] == user_data["name"]
        assert created_user["email"] == user_data["email"]
        assert "password" not in created_user  # Should not return password

    def test_authentication(self):
        """Test JWT authentication."""
        # First create a user
        user_data = {"email": "test@example.com", "password": "secure_password"}

        # Login
        response = client.post("/auth/login", json=user_data)
        assert response.status_code == 200

        token = response.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}"}

        # Access protected endpoint
        response = client.get("/users/me", headers=headers)
        assert response.status_code == 200
            ''',

            'integration_testing': '''
import pytest
import asyncio
from httpx import AsyncClient
from app.main import app
from app.database import init_db, cleanup_db

@pytest.fixture(scope="session")
def event_loop():
    """Create event loop for async tests."""
    loop = asyncio.get_event_loop_policy().new_event_loop()
    yield loop
    loop.close()

@pytest.fixture(scope="session")
async def setup_database():
    """Setup test database."""
    await init_db()
    yield
    await cleanup_db()

@pytest.mark.asyncio
async def test_full_user_workflow(setup_database):
    """Test complete user workflow."""
    async with AsyncClient(app=app, base_url="http://test") as client:
        # Register user
        user_data = {
            "name": "Integration Test User",
            "email": "integration@example.com",
            "password": "test_password"
        }

        response = await client.post("/register", json=user_data)
        assert response.status_code == 201

        # Login
        login_data = {"email": user_data["email"], "password": user_data["password"]}
        response = await client.post("/login", json=login_data)
        assert response.status_code == 200

        token = response.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}"}

        # Update profile
        update_data = {"name": "Updated Name"}
        response = await client.patch("/profile", json=update_data, headers=headers)
        assert response.status_code == 200
        assert response.json()["name"] == "Updated Name"
            '''
        }

        for test_type, code in testing_examples.items():
            if any(term in query.lower() for term in test_type.split('_')):
                snippets.append(code.strip())

        return snippets

    def fetch_client_integration_examples(self, query: str, logger) -> List[str]:
        """Fetch examples for common client integrations."""
        logger.info(f"Fetching client integration examples for: {query}")
        snippets = []

        integration_examples = {
            'stripe': '''
import stripe
from fastapi import HTTPException

stripe.api_key = "sk_test_..."

class PaymentService:
    @staticmethod
    def create_payment_intent(amount: int, currency: str = "usd"):
        """Create Stripe payment intent."""
        try:
            intent = stripe.PaymentIntent.create(
                amount=amount * 100,  # Convert to cents
                currency=currency,
                metadata={'integration_check': 'accept_a_payment'}
            )
            return intent
        except stripe.error.StripeError as e:
            raise HTTPException(status_code=400, detail=str(e))

    @staticmethod
    def create_customer(email: str, name: str):
        """Create Stripe customer."""
        try:
            customer = stripe.Customer.create(
                email=email,
                name=name
            )
            return customer
        except stripe.error.StripeError as e:
            raise HTTPException(status_code=400, detail=str(e))
            ''',

            'twilio': '''
from twilio.rest import Client
import os

class SMSService:
    def __init__(self):
        self.account_sid = os.getenv('TWILIO_ACCOUNT_SID')
        self.auth_token = os.getenv('TWILIO_AUTH_TOKEN')
        self.from_number = os.getenv('TWILIO_PHONE_NUMBER')
        self.client = Client(self.account_sid, self.auth_token)

    def send_sms(self, to_number: str, message: str):
        """Send SMS via Twilio."""
        try:
            message = self.client.messages.create(
                body=message,
                from_=self.from_number,
                to=to_number
            )
            return message.sid
        except Exception as e:
            raise Exception(f"Failed to send SMS: {str(e)}")

    def send_verification_code(self, phone_number: str):
        """Send verification code."""
        import random
        code = random.randint(100000, 999999)
        message = f"Your verification code is: {code}"

        message_sid = self.send_sms(phone_number, message)
        return {'code': code, 'message_sid': message_sid}
            ''',

            'sendgrid': '''
import sendgrid
from sendgrid.helpers.mail import Mail, Email, To, Content
import os

class EmailService:
    def __init__(self):
        self.sg = sendgrid.SendGridAPIClient(api_key=os.getenv('SENDGRID_API_KEY'))

    def send_email(self, to_email: str, subject: str, html_content: str):
        """Send email via SendGrid."""
        from_email = Email("noreply@yourapp.com")
        to_email = To(to_email)
        content = Content("text/html", html_content)

        mail = Mail(from_email, to_email, subject, content)

        try:
            response = self.sg.client.mail.send.post(request_body=mail.get())
            return response.status_code
        except Exception as e:
            raise Exception(f"Failed to send email: {str(e)}")

    def send_template_email(self, to_email: str, template_id: str, dynamic_data: dict):
        """Send templated email."""
        message = Mail(
            from_email='noreply@yourapp.com',
            to_emails=to_email
        )
        message.template_id = template_id
        message.dynamic_template_data = dynamic_data

        try:
            response = self.sg.send(message)
            return response.status_code
        except Exception as e:
            raise Exception(f"Failed to send template email: {str(e)}")
            ''',

            'aws': '''
import boto3
from botocore.exceptions import ClientError
import os
import json

class AWSService:
    def __init__(self):
        self.s3_client = boto3.client('s3')
        self.ses_client = boto3.client('ses')
        self.lambda_client = boto3.client('lambda')

    def upload_to_s3(self, file_path: str, bucket: str, key: str):
        """Upload file to S3."""
        try:
            self.s3_client.upload_file(file_path, bucket, key)
            return f"https://{bucket}.s3.amazonaws.com/{key}"
        except ClientError as e:
            raise Exception(f"S3 upload failed: {str(e)}")

    def send_ses_email(self, to_email: str, subject: str, body: str):
        """Send email via SES."""
        try:
            response = self.ses_client.send_email(
                Source='noreply@yourapp.com',
                Destination={'ToAddresses': [to_email]},
                Message={
                    'Subject': {'Data': subject},
                    'Body': {'Html': {'Data': body}}
                }
            )
            return response['MessageId']
        except ClientError as e:
            raise Exception(f"SES email failed: {str(e)}")

    def invoke_lambda(self, function_name: str, payload: dict):
        """Invoke Lambda function."""
        try:
            response = self.lambda_client.invoke(
                FunctionName=function_name,
                Payload=json.dumps(payload)
            )
            return response['Payload'].read()
        except ClientError as e:
            raise Exception(f"Lambda invocation failed: {str(e)}")
            '''
        }

        # Check if query mentions any integrations
        for service, code in integration_examples.items():
            if service in query.lower():
                snippets.append(code.strip())

        return snippets

    def fetch_deployment_examples(self, query: str, logger) -> List[str]:
        """Fetch deployment and DevOps examples."""
        logger.info(f"Fetching deployment examples for: {query}")
        snippets = []

        deployment_terms = ['deploy', 'docker', 'kubernetes', 'heroku', 'aws', 'production']
        if not any(term in query.lower() for term in deployment_terms):
            return snippets

        deployment_examples = {
            'dockerfile': '''
# Dockerfile for FastAPI application
FROM python:3.11-slim

WORKDIR /app

# Install system dependencies
RUN apt-get update && apt-get install -y \\
    gcc \\
    && rm -rf /var/lib/apt/lists/*

# Copy requirements first for better caching
COPY requirements.txt .
RUN pip install --no-cache-dir -r requirements.txt

# Copy application code
COPY . .

# Create non-root user
RUN adduser --disabled-password --gecos '' appuser
RUN chown -R appuser:appuser /app
USER appuser

# Expose port
EXPOSE 8000

# Health check
HEALTHCHECK --interval=30s --timeout=30s --start-period=5s --retries=3 \\
    CMD curl -f http://localhost:8000/health || exit 1

# Run application
CMD ["uvicorn", "app.main:app", "--host", "0.0.0.0", "--port", "8000"]
            ''',

            'docker_compose': '''
# docker-compose.yml for full stack application
version: '3.8'

services:
  web:
    build: .
    ports:
      - "8000:8000"
    environment:
      - DATABASE_URL=postgresql://user:password@db:5432/appdb
      - REDIS_URL=redis://redis:6379
    depends_on:
      - db
      - redis
    volumes:
      - ./app:/app/app
    command: uvicorn app.main:app --host 0.0.0.0 --port 8000 --reload

  db:
    image: postgres:13
    environment:
      - POSTGRES_USER=user
      - POSTGRES_PASSWORD=password
      - POSTGRES_DB=appdb
    volumes:
      - postgres_data:/var/lib/postgresql/data
    ports:
      - "5432:5432"

  redis:
    image: redis:6-alpine
    ports:
      - "6379:6379"

  worker:
    build: .
    command: celery -A app.celery worker --loglevel=info
    environment:
      - DATABASE_URL=postgresql://user:password@db:5432/appdb
      - REDIS_URL=redis://redis:6379
    depends_on:
      - db
      - redis

volumes:
  postgres_data:
            ''',

            'github_actions': '''
# .github/workflows/ci.yml
name: CI/CD Pipeline

on:
  push:
    branches: [ main, develop ]
  pull_request:
    branches: [ main ]

jobs:
  test:
    runs-on: ubuntu-latest

    services:
      postgres:
        image: postgres:13
        env:
          POSTGRES_PASSWORD: postgres
          POSTGRES_DB: test_db
        options: >-
          --health-cmd pg_isready
          --health-interval 10s
          --health-timeout 5s
          --health-retries 5

    steps:
    - uses: actions/checkout@v3

    - name: Set up Python
      uses: actions/setup-python@v3
      with:
        python-version: '3.11'

    - name: Install dependencies
      run: |
        python -m pip install --upgrade pip
        pip install -r requirements.txt
        pip install pytest pytest-cov

    - name: Run tests
      run: |
        pytest --cov=app --cov-report=xml

    - name: Upload coverage
      uses: codecov/codecov-action@v3

  deploy:
    needs: test
    runs-on: ubuntu-latest
    if: github.ref == 'refs/heads/main'

    steps:
    - uses: actions/checkout@v3

    - name: Deploy to production
      run: |
        echo "Deploy to production server"
        # Add your deployment commands here
            '''
        }

        for deployment_type, code in deployment_examples.items():
            if any(term in query.lower() for term in deployment_type.split('_')):
                snippets.append(code.strip())

        return snippets

    def fetch_data_processing_examples(self, query: str, logger) -> List[str]:
        """Fetch data processing examples for freelance work."""
        logger.info(f"Fetching data processing examples for: {query}")
        snippets = []

        data_terms = ['data', 'pandas', 'csv', 'excel', 'etl', 'process']
        if not any(term in query.lower() for term in data_terms):
            return snippets

        data_examples = {
            'csv_processing': '''
import pandas as pd
import numpy as np
from pathlib import Path

class CSVProcessor:
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.df = None

    def load_data(self):
        """Load CSV with error handling."""
        try:
            self.df = pd.read_csv(self.file_path)
            print(f"Loaded {len(self.df)} rows from {self.file_path}")
            return self.df
        except Exception as e:
            print(f"Error loading CSV: {e}")
            return None

    def clean_data(self):
        """Basic data cleaning operations."""
        if self.df is None:
            return None

        # Remove duplicates
        before_count = len(self.df)
        self.df = self.df.drop_duplicates()
        after_count = len(self.df)
        print(f"Removed {before_count - after_count} duplicates")

        # Handle missing values
        self.df = self.df.fillna(method='ffill')

        # Clean string columns
        string_cols = self.df.select_dtypes(include=['object']).columns
        for col in string_cols:
            self.df[col] = self.df[col].str.strip()

        return self.df

    def export_cleaned_data(self, output_path: str):
        """Export cleaned data to new CSV."""
        if self.df is not None:
            self.df.to_csv(output_path, index=False)
            print(f"Exported cleaned data to {output_path}")
            ''',

            'excel_automation': '''
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference

class ExcelReportGenerator:
    def __init__(self):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active

    def create_sales_report(self, data: pd.DataFrame):
        """Create formatted sales report."""
        # Write data to worksheet
        for r_idx, row in enumerate(data.itertuples(index=False), 1):
            for c_idx, value in enumerate(row, 1):
                self.worksheet.cell(row=r_idx, column=c_idx, value=value)

        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for col in range(1, len(data.columns) + 1):
            cell = self.worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

        # Add chart
        chart = BarChart()
        chart.title = "Sales by Category"

        # Assuming data has sales values in column 2
        data_ref = Reference(self.worksheet, min_col=2, min_row=2, max_row=len(data) + 1)
        cats_ref = Reference(self.worksheet, min_col=1, min_row=2, max_row=len(data) + 1)

        chart.add_data(data_ref)
        chart.set_categories(cats_ref)

        self.worksheet.add_chart(chart, "E2")

    def save_report(self, filename: str):
        """Save the Excel report."""
        self.workbook.save(filename)
        print(f"Report saved as {filename}")
            '''
        }

        for data_type, code in data_examples.items():
            if any(term in query.lower() for term in data_type.split('_')):
                snippets.append(code.strip())

        return snippets


def search_freelancer_sources(query: str, logger) -> List[str]:
    """Search freelancer-specific Python sources."""
    try:
        from config import USER_AGENT, DEFAULT_REQUEST_TIMEOUT
    except ImportError:
        USER_AGENT = "RAGContentScraper/1.0"
        DEFAULT_REQUEST_TIMEOUT = 15

    freelancer_sources = FreelancerPythonSources(USER_AGENT, DEFAULT_REQUEST_TIMEOUT)
    all_snippets = []

    # Fetch different types of examples
    source_methods = [
        freelancer_sources.fetch_python_patterns,
        freelancer_sources.fetch_fastapi_examples,
        freelancer_sources.fetch_automation_scripts,
        freelancer_sources.fetch_testing_patterns,
        freelancer_sources.fetch_client_integration_examples,
        freelancer_sources.fetch_deployment_examples,
        freelancer_sources.fetch_data_processing_examples
    ]

    for method in source_methods:
        try:
            snippets = method(query, logger)
            all_snippets.extend(snippets)
        except Exception as e:
            logger.error(f"Error in {method.__name__}: {e}")

    return all_snippets